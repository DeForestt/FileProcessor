from re import I, S
from os.path import exists
from openpyxl.worksheet import worksheet
import pandas
from openpyxl import load_workbook

def extractColorList(sh : worksheet):
    colors = []
    for index, row in enumerate(sh.iter_rows()):
        if index != 0:
            color_in_hex = str(row[0].fill.start_color.index) # this gives you Hexadecimal value of the color
            colors.append(color_in_hex)
    return colors

def genDic(colors : list):
    result = dict()

    i = 0
    for color in colors:
        if str(color) not in result.keys():
            result[str(color)] = i
            i += 1
    return result

#create an address record
def create(strin: list, index):
    map(str.strip, strin)
    if len(strin) == 3:
        if ' ' in strin[2]:
            spl = strin[2].split()
            return [strin[0], strin[1], '', spl[0], '', spl[1], 1]
    elif len(strin) == 4:
        if ' ' in strin[3]:
            spl = strin[3].split()
            return [strin[0], strin[1], strin[2], strin[3], spl[0], spl[1], 1]
    elif len(strin) == 5:
        return [strin[0], strin[1], '' ,strin[2], strin[3], strin[4], 1]
    elif len(strin) == 6:
        return trp(strin, 7)

    ret = trp(strin, 7)
    ret[6] = 1
    return ret

def trp(l, n):
    """ Truncate or pad a list """
    r = l[:n]
    if len(r) < n:
        r.extend([0] * (n - len(r)))
    return r

def lumenProc():
    print("Starting Processing Lumen Report...")
    xl = 'LumenInput.xlsx'
    wb = load_workbook(xl, data_only = True)
    sh = wb['Worksheet']

    colors = extractColorList(sh)
    colorDict = genDic(colors)

    df = pandas.read_excel(xl, sheet_name='Worksheet', index_col=None, header = 0)
    #rename df header
    df.set_axis(['Account No','Customer Name','Amount', 'Address', 'CONTACT NAME', 'Dunning Email', 'Dunning Phone', 'Dunning ', 'Dunning .1', 'Dunning .2'], axis=1, inplace=True)
    addresses = pandas.DataFrame(columns=["Street", "City", "County", "State", "Country", "Zip", "Review"])

    addressPut = list()
    for i in range(0, len(colorDict)):
        nframe = pandas.DataFrame(columns=["Street", "City", "County", "State", "Country", "Zip", "Review"])
        addressPut.append(nframe)
    origPut = list()
    for i in range(0, len(colorDict)):
        nframe = pandas.DataFrame(columns=df.columns)
        origPut.append(nframe)
    
    for index, row in df.iterrows():

        bin = create(str(row['Address']).split(','), index)

        si = len(addresses)
        addresses.loc[si] = bin

        listIndex = colorDict[str(colors[index])]
        
        si = len(addressPut[listIndex])
        addressPut[listIndex].loc[si] = bin

        si = len(origPut[listIndex])
        origPut[listIndex].loc[si] = row


    writer = pandas.ExcelWriter('LumenOutput.xlsx', engine='xlsxwriter')

    for index, frame in enumerate(origPut):
        out = pandas.concat([frame[['Account No','Customer Name','Amount', 'Address']], addressPut[index], frame[['CONTACT NAME', 'Dunning Email', 'Dunning Phone', 'Dunning ', 'Dunning .1', 'Dunning .2']]], axis=1)
        out.to_excel(writer, sheet_name='Coustomer List-{}'.format(index), index=False)
    writer.save()
    print("Lumen Report Processed Sucsessfully")
